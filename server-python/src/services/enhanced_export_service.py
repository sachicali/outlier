import os
import io
import json
import uuid
import csv
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, Union
from urllib.parse import quote

# Excel and data processing
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    import pandas as pd
except ImportError:
    openpyxl = None
    pd = None

# PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.barcharts import VerticalBarChart
except ImportError:
    pass

# Charting
try:
    import matplotlib.pyplot as plt
    import seaborn as sns
    import matplotlib
    matplotlib.use('Agg')  # Use non-interactive backend
except ImportError:
    plt = None
    sns = None

from utils.logger import logger
from db import redis_client, get_db_session
from models.export_job import ExportJob, ExportStatus, ExportFormat
from repositories.export_repository import ExportRepository
from .export_service import ExportService  # Base service

class EnhancedExportService(ExportService):
    """Enhanced export service with database persistence and advanced features"""
    
    def __init__(self):
        super().__init__()
        self.export_repository = None
    
    def _get_repository(self) -> ExportRepository:
        """Get export repository with database session"""
        if not self.export_repository:
            session = get_db_session()
            self.export_repository = ExportRepository(session)
        return self.export_repository
    
    def create_export_job_db(
        self, 
        analysis_id: str, 
        format_type: str, 
        user_id: int,
        export_config: Optional[Dict[str, Any]] = None
    ) -> str:
        """Create export job with database persistence"""
        job_id = str(uuid.uuid4())
        
        try:
            repo = self._get_repository()
            
            # Create job in database
            job = repo.create_export_job(
                job_id=job_id,
                user_id=user_id,
                analysis_id=analysis_id,
                format_type=format_type,
                export_config=export_config or {}
            )
            
            # Also store in Redis for backward compatibility
            job_data = {
                'id': job_id,
                'analysis_id': analysis_id,
                'format': format_type,
                'user_id': user_id,
                'status': 'pending',
                'created_at': datetime.utcnow().isoformat(),
                'progress': 0
            }
            redis_client.setex(f'export_job:{job_id}', 3600, json.dumps(job_data))
            
            logger.info(f"Created enhanced export job {job_id} for analysis {analysis_id}")
            return job_id
            
        except Exception as e:
            logger.error(f"Failed to create export job: {e}")
            raise
    
    def get_export_job_status_db(self, job_id: str) -> Optional[Dict[str, Any]]:
        """Get export job status from database"""
        try:
            repo = self._get_repository()
            job = repo.get_job_by_id(job_id)
            
            if job:
                return job.to_dict()
            
            # Fallback to Redis
            return super().get_export_job_status(job_id)
            
        except Exception as e:
            logger.error(f"Error getting job status from DB: {e}")
            # Fallback to Redis
            return super().get_export_job_status(job_id)
    
    def update_job_progress_db(
        self, 
        job_id: str, 
        progress: int, 
        message: Optional[str] = None
    ):
        """Update job progress in database"""
        try:
            repo = self._get_repository()
            repo.update_job_progress(job_id, progress, message)
            
            # Also update Redis for real-time updates
            job_data = redis_client.get(f'export_job:{job_id}')
            if job_data:
                job = json.loads(job_data)
                job['progress'] = progress
                if message:
                    job['message'] = message
                redis_client.setex(f'export_job:{job_id}', 3600, json.dumps(job))
                
        except Exception as e:
            logger.error(f"Error updating job progress in DB: {e}")
    
    def mark_job_completed_db(
        self, 
        job_id: str, 
        file_path: str, 
        file_size: Optional[int] = None,
        filename: Optional[str] = None
    ):
        """Mark job as completed in database"""
        try:
            repo = self._get_repository()
            
            # Determine MIME type based on file extension
            mime_types = {
                '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                '.csv': 'text/csv',
                '.pdf': 'application/pdf',
                '.json': 'application/json',
                '.html': 'text/html'
            }
            
            file_ext = os.path.splitext(file_path)[1].lower()
            mime_type = mime_types.get(file_ext, 'application/octet-stream')
            
            repo.mark_job_completed(
                job_id=job_id,
                file_path=file_path,
                file_size=file_size,
                filename=filename,
                mime_type=mime_type
            )
            
            # Update Redis
            job_data = redis_client.get(f'export_job:{job_id}')
            if job_data:
                job = json.loads(job_data)
                job['status'] = 'completed'
                job['progress'] = 100
                job['file_path'] = file_path
                job['filename'] = filename
                if file_size:
                    job['file_size'] = file_size
                redis_client.setex(f'export_job:{job_id}', 86400, json.dumps(job))
                
        except Exception as e:
            logger.error(f"Error marking job completed in DB: {e}")
    
    def mark_job_failed_db(self, job_id: str, error_message: str):
        """Mark job as failed in database"""
        try:
            repo = self._get_repository()
            repo.mark_job_failed(job_id, error_message)
            
            # Update Redis
            job_data = redis_client.get(f'export_job:{job_id}')
            if job_data:
                job = json.loads(job_data)
                job['status'] = 'failed'
                job['error'] = error_message
                redis_client.setex(f'export_job:{job_id}', 3600, json.dumps(job))
                
        except Exception as e:
            logger.error(f"Error marking job failed in DB: {e}")
    
    def get_user_export_jobs(
        self, 
        user_id: int, 
        limit: int = 50, 
        status: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """Get user's export jobs from database"""
        try:
            repo = self._get_repository()
            jobs = repo.get_user_jobs(user_id, limit, status)
            return [job.to_dict() for job in jobs]
            
        except Exception as e:
            logger.error(f"Error getting user export jobs: {e}")
            return []
    
    def cancel_export_job(self, job_id: str, user_id: int) -> bool:
        """Cancel an export job"""
        try:
            repo = self._get_repository()
            job = repo.cancel_job(job_id, user_id)
            
            if job:
                # Update Redis
                job_data = redis_client.get(f'export_job:{job_id}')
                if job_data:
                    job_dict = json.loads(job_data)
                    job_dict['status'] = 'cancelled'
                    redis_client.setex(f'export_job:{job_id}', 3600, json.dumps(job_dict))
                
                logger.info(f"Cancelled export job {job_id} for user {user_id}")
                return True
            
            return False
            
        except Exception as e:
            logger.error(f"Error cancelling export job: {e}")
            return False
    
    def retry_export_job(self, job_id: str, user_id: int) -> bool:
        """Retry a failed export job"""
        try:
            repo = self._get_repository()
            job = repo.get_job_by_id(job_id)
            
            if not job or job.user_id != user_id:
                return False
            
            if job.can_retry():
                retry_job = repo.retry_job(job_id)
                if retry_job:
                    # Update Redis
                    job_data = {
                        'id': job_id,
                        'analysis_id': job.analysis_id,
                        'format': job.format,
                        'user_id': user_id,
                        'status': 'pending',
                        'progress': 0
                    }
                    redis_client.setex(f'export_job:{job_id}', 3600, json.dumps(job_data))
                    
                    logger.info(f"Retrying export job {job_id} for user {user_id}")
                    return True
            
            return False
            
        except Exception as e:
            logger.error(f"Error retrying export job: {e}")
            return False
    
    def cleanup_old_exports(self, max_age_hours: int = 24) -> int:
        """Clean up old export files and jobs"""
        try:
            repo = self._get_repository()
            cleaned_count = repo.cleanup_expired_jobs(delete_files=True)
            
            # Also clean up old Redis entries
            super().cleanup_temp_files(max_age_hours)
            
            return cleaned_count
            
        except Exception as e:
            logger.error(f"Error during enhanced cleanup: {e}")
            return 0
    
    def get_export_statistics(self, user_id: Optional[int] = None, days: int = 30) -> Dict[str, Any]:
        """Get export statistics"""
        try:
            repo = self._get_repository()
            return repo.get_job_statistics(user_id, days)
            
        except Exception as e:
            logger.error(f"Error getting export statistics: {e}")
            return {'error': str(e)}
    
    def get_user_export_quota(self, user_id: int, period_hours: int = 24) -> Dict[str, Any]:
        """Get user's export quota usage"""
        try:
            repo = self._get_repository()
            return repo.get_user_export_quota(user_id, period_hours)
            
        except Exception as e:
            logger.error(f"Error getting user export quota: {e}")
            return {'error': str(e)}
    
    def generate_advanced_excel_export(
        self, 
        analysis_id: str, 
        user_id: int = None,
        include_charts: bool = True,
        include_summary: bool = True,
        include_recommendations: bool = True
    ) -> io.BytesIO:
        """Generate advanced Excel export with enhanced features"""
        if not openpyxl or not pd:
            raise ValueError("Advanced Excel export requires openpyxl and pandas")
        
        data = self.get_analysis_data(analysis_id)
        if not data:
            raise ValueError(f"Analysis {analysis_id} not found")
        
        # Create workbook with professional styling
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Company/Brand styling
        brand_color = '2E75B6'
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color=brand_color, end_color=brand_color, fill_type='solid')
        
        if include_summary:
            self._create_executive_summary_sheet(wb, data, header_font, header_fill)
        
        # Enhanced outlier results with advanced filtering
        self._create_enhanced_outliers_sheet(wb, data, header_font, header_fill)
        
        # Channel performance analysis
        self._create_channel_performance_sheet(wb, data, header_font, header_fill)
        
        if include_charts:
            self._create_advanced_charts_sheet(wb, data)
        
        if include_recommendations:
            self._create_recommendations_sheet(wb, data, header_font, header_fill)
        
        # Data validation and pivot tables
        self._create_data_analysis_sheet(wb, data, header_font, header_fill)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Advanced Excel export completed for analysis {analysis_id}")
        return output
    
    def _create_executive_summary_sheet(self, wb, data, header_font, header_fill):
        """Create executive summary sheet with key insights"""
        ws = wb.create_sheet("Executive Summary")
        
        # Title and branding
        ws['A1'] = 'YouTube Outlier Discovery - Executive Summary'
        ws['A1'].font = Font(name='Arial', size=18, bold=True, color='2E75B6')
        ws.merge_cells('A1:E1')
        
        # Key metrics dashboard
        row = 3
        metrics = [
            ('Total Outliers Discovered', data['summary'].get('totalOutliers', 0)),
            ('Channels Analyzed', data['summary'].get('channelsAnalyzed', 0)),
            ('Average Outlier Score', self._calculate_average_outlier_score(data)),
            ('Top Performing Channel', self._get_top_channel(data)),
            ('Recommended Focus Areas', self._get_focus_areas(data))
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Quick insights
        ws[f'A{row + 1}'] = 'Key Insights:'
        ws[f'A{row + 1}'].font = Font(bold=True, size=14)
        
        insights = self._generate_insights(data)
        for i, insight in enumerate(insights, row + 2):
            ws[f'A{i}'] = f"• {insight}"
    
    def _create_enhanced_outliers_sheet(self, wb, data, header_font, header_fill):
        """Create enhanced outliers sheet with conditional formatting"""
        ws = wb.create_sheet("Detailed Results")
        
        # Enhanced headers with additional metrics
        headers = [
            'Rank', 'Video Title', 'Channel Name', 'Video URL', 'Channel URL',
            'Outlier Score', 'Brand Fit Score', 'Performance Rating', 'Views', 'Likes', 'Comments',
            'Engagement Rate', 'Subscriber Count', 'Video Count', 'Published Date', 
            'Days Since Published', 'Content Category', 'Estimated Revenue',
            'Competition Level', 'Opportunity Score', 'Tags', 'Description Preview'
        ]
        
        # Set headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Enhanced data with calculated fields
        results = data['results']
        for row, result in enumerate(results, 2):
            # Calculate additional metrics
            views = int(result.get('statistics', {}).get('viewCount', 0))
            likes = int(result.get('statistics', {}).get('likeCount', 0))
            comments = int(result.get('statistics', {}).get('commentCount', 0))
            
            engagement_rate = ((likes + comments) / max(views, 1)) * 100
            performance_rating = self._calculate_performance_rating(result)
            opportunity_score = self._calculate_opportunity_score(result)
            
            # Populate row with enhanced data
            row_data = [
                row - 1,  # Rank
                result.get('snippet', {}).get('title', 'N/A'),
                result.get('channelInfo', {}).get('snippet', {}).get('title', 'N/A'),
                f"https://www.youtube.com/watch?v={result.get('id', {}).get('videoId', result.get('id', ''))}",
                f"https://www.youtube.com/channel/{result.get('channelInfo', {}).get('id', '')}",
                round(result.get('outlierScore', 0), 2),
                round(result.get('brandFit', 0), 2),
                performance_rating,
                views,
                likes,
                comments,
                round(engagement_rate, 3),
                int(result.get('channelInfo', {}).get('statistics', {}).get('subscriberCount', 0)),
                int(result.get('channelInfo', {}).get('statistics', {}).get('videoCount', 0)),
                result.get('snippet', {}).get('publishedAt', 'N/A'),
                self._calculate_days_since_published(result),
                self._categorize_content(result),
                self._estimate_revenue(result),
                self._assess_competition_level(result),
                opportunity_score,
                ', '.join(result.get('snippet', {}).get('tags', [])[:5]),
                result.get('snippet', {}).get('description', '')[:100] + '...'
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Add conditional formatting
        self._add_conditional_formatting(ws, len(results))
    
    def _create_channel_performance_sheet(self, wb, data, header_font, header_fill):
        """Create channel performance analysis sheet"""
        ws = wb.create_sheet("Channel Performance")
        
        # Group and analyze by channel
        channel_analysis = self._analyze_channels(data)
        
        headers = [
            'Channel Name', 'Total Outliers', 'Avg Outlier Score', 'Avg Brand Fit',
            'Total Views', 'Subscriber Count', 'Success Rate', 'Content Consistency',
            'Growth Potential', 'Recommendation'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row, (channel_id, analysis) in enumerate(channel_analysis.items(), 2):
            row_data = [
                analysis['name'],
                analysis['outlier_count'],
                round(analysis['avg_outlier_score'], 2),
                round(analysis['avg_brand_fit'], 2),
                analysis['total_views'],
                analysis['subscriber_count'],
                f"{analysis['success_rate']:.1f}%",
                analysis['consistency_score'],
                analysis['growth_potential'],
                analysis['recommendation']
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
    
    def _create_advanced_charts_sheet(self, wb, data):
        """Create advanced charts and visualizations"""
        ws = wb.create_sheet("Analytics & Charts")
        
        # Performance distribution chart
        ws['A1'] = 'Performance Distribution Analysis'
        ws['A1'].font = Font(bold=True, size=14)
        
        # Create data for charts
        results = data['results'][:20]  # Top 20 for charts
        
        # Outlier Score vs Brand Fit scatter
        ws['A3'] = 'Video Title'
        ws['B3'] = 'Outlier Score'
        ws['C3'] = 'Brand Fit Score'
        ws['D3'] = 'Views'
        
        for i, result in enumerate(results, 4):
            ws[f'A{i}'] = result.get('snippet', {}).get('title', 'N/A')[:30]
            ws[f'B{i}'] = result.get('outlierScore', 0)
            ws[f'C{i}'] = result.get('brandFit', 0)
            ws[f'D{i}'] = int(result.get('statistics', {}).get('viewCount', 0))
        
        # Create scatter chart
        chart = openpyxl.chart.ScatterChart()
        chart.title = "Outlier Score vs Brand Fit Analysis"
        chart.style = 13
        chart.x_axis.title = 'Outlier Score'
        chart.y_axis.title = 'Brand Fit Score'
        
        xvalues = Reference(ws, min_col=2, min_row=4, max_row=len(results) + 3)
        yvalues = Reference(ws, min_col=3, min_row=4, max_row=len(results) + 3)
        series = openpyxl.chart.Series(yvalues, xvalues, title="Videos")
        chart.series.append(series)
        
        ws.add_chart(chart, "F3")
    
    def _create_recommendations_sheet(self, wb, data, header_font, header_fill):
        """Create actionable recommendations sheet"""
        ws = wb.create_sheet("Recommendations")
        
        ws['A1'] = 'Strategic Recommendations'
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='2E75B6')
        
        recommendations = self._generate_recommendations(data)
        
        row = 3
        for category, recs in recommendations.items():
            ws[f'A{row}'] = category
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            for rec in recs:
                ws[f'A{row}'] = f"• {rec}"
                row += 1
            
            row += 1  # Add spacing
    
    def _create_data_analysis_sheet(self, wb, data, header_font, header_fill):
        """Create data analysis sheet with pivot table data"""
        ws = wb.create_sheet("Data Analysis")
        
        # Summary statistics
        ws['A1'] = 'Statistical Analysis'
        ws['A1'].font = Font(bold=True, size=14)
        
        stats = self._calculate_statistics(data)
        
        row = 3
        for stat_name, value in stats.items():
            ws[f'A{row}'] = stat_name
            ws[f'B{row}'] = value
            row += 1
    
    # Helper methods for advanced analysis
    def _calculate_average_outlier_score(self, data):
        results = data['results']
        if not results:
            return 0
        return sum(r.get('outlierScore', 0) for r in results) / len(results)
    
    def _get_top_channel(self, data):
        channel_scores = {}
        for result in data['results']:
            channel = result.get('channelInfo', {}).get('snippet', {}).get('title', 'Unknown')
            score = result.get('outlierScore', 0)
            if channel not in channel_scores:
                channel_scores[channel] = []
            channel_scores[channel].append(score)
        
        if channel_scores:
            top_channel = max(channel_scores.items(), key=lambda x: sum(x[1])/len(x[1]))
            return top_channel[0]
        return 'N/A'
    
    def _get_focus_areas(self, data):
        # Analyze content patterns to suggest focus areas
        content_categories = {}
        for result in data['results']:
            category = self._categorize_content(result)
            content_categories[category] = content_categories.get(category, 0) + 1
        
        if content_categories:
            top_category = max(content_categories.items(), key=lambda x: x[1])
            return top_category[0]
        return 'Gaming Content'
    
    def _generate_insights(self, data):
        insights = []
        results = data['results']
        
        if results:
            avg_score = sum(r.get('outlierScore', 0) for r in results) / len(results)
            insights.append(f"Average outlier score of {avg_score:.1f} indicates strong content opportunities")
            
            high_brand_fit = len([r for r in results if r.get('brandFit', 0) > 7])
            insights.append(f"{high_brand_fit} videos show high brand compatibility (>7.0 score)")
            
            recent_videos = len([r for r in results if self._is_recent_video(r)])
            insights.append(f"{recent_videos} outliers are recent content (last 7 days)")
        
        return insights
    
    def _calculate_performance_rating(self, result):
        outlier_score = result.get('outlierScore', 0)
        brand_fit = result.get('brandFit', 0)
        
        combined_score = (outlier_score * 0.7) + (brand_fit * 0.3)
        
        if combined_score >= 25:
            return 'Excellent'
        elif combined_score >= 20:
            return 'Very Good'
        elif combined_score >= 15:
            return 'Good'
        elif combined_score >= 10:
            return 'Fair'
        else:
            return 'Poor'
    
    def _calculate_opportunity_score(self, result):
        # Complex scoring algorithm considering multiple factors
        outlier_score = result.get('outlierScore', 0)
        brand_fit = result.get('brandFit', 0)
        views = int(result.get('statistics', {}).get('viewCount', 0))
        subscribers = int(result.get('channelInfo', {}).get('statistics', {}).get('subscriberCount', 1))
        
        # Normalize and weight different factors
        score = (outlier_score * 0.4) + (brand_fit * 0.3) + min((views / subscribers) * 10, 20) * 0.3
        
        return round(score, 2)
    
    def _calculate_days_since_published(self, result):
        published_at = result.get('snippet', {}).get('publishedAt')
        if published_at:
            try:
                pub_date = datetime.fromisoformat(published_at.replace('Z', '+00:00'))
                return (datetime.now(pub_date.tzinfo) - pub_date).days
            except:
                pass
        return 0
    
    def _categorize_content(self, result):
        title = result.get('snippet', {}).get('title', '').lower()
        tags = [tag.lower() for tag in result.get('snippet', {}).get('tags', [])]
        
        categories = {
            'Gaming': ['game', 'gaming', 'minecraft', 'fortnite', 'roblox', 'valorant'],
            'Entertainment': ['funny', 'comedy', 'entertainment', 'reaction', 'challenge'],
            'Educational': ['tutorial', 'how to', 'guide', 'learn', 'education'],
            'Technology': ['tech', 'review', 'unboxing', 'gadget', 'phone'],
            'Lifestyle': ['vlog', 'daily', 'life', 'routine', 'lifestyle']
        }
        
        for category, keywords in categories.items():
            if any(keyword in title or keyword in ' '.join(tags) for keyword in keywords):
                return category
        
        return 'Other'
    
    def _estimate_revenue(self, result):
        views = int(result.get('statistics', {}).get('viewCount', 0))
        # Rough estimate: $1-3 per 1000 views
        estimated = (views / 1000) * 2
        return f"${estimated:.0f}"
    
    def _assess_competition_level(self, result):
        outlier_score = result.get('outlierScore', 0)
        
        if outlier_score >= 30:
            return 'Low'
        elif outlier_score >= 20:
            return 'Medium'
        else:
            return 'High'
    
    def _is_recent_video(self, result):
        return self._calculate_days_since_published(result) <= 7
    
    def _add_conditional_formatting(self, ws, num_results):
        """Add conditional formatting to highlight important data"""
        from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
        from openpyxl.styles import PatternFill
        
        # Outlier score color scale (column F)
        outlier_col = f'F2:F{num_results + 1}'
        ws.conditional_formatting.add(outlier_col,
                                    ColorScaleRule(start_type='min', start_color='FFFF00',
                                                 end_type='max', end_color='FF0000'))
        
        # Brand fit color scale (column G)
        brand_fit_col = f'G2:G{num_results + 1}'
        ws.conditional_formatting.add(brand_fit_col,
                                    ColorScaleRule(start_type='min', start_color='FFE6E6',
                                                 end_type='max', end_color='90EE90'))
        
        # Highlight excellent performance ratings
        performance_col = f'H2:H{num_results + 1}'
        ws.conditional_formatting.add(performance_col,
                                    CellIsRule(operator='equal',
                                             formula=['"Excellent"'],
                                             fill=PatternFill(start_color='90EE90', end_color='90EE90')))
    
    def _analyze_channels(self, data):
        """Analyze channels and return performance metrics"""
        channel_data = {}
        
        for result in data['results']:
            channel_info = result.get('channelInfo', {})
            channel_id = channel_info.get('id', 'unknown')
            
            if channel_id not in channel_data:
                channel_data[channel_id] = {
                    'name': channel_info.get('snippet', {}).get('title', 'Unknown'),
                    'outliers': [],
                    'subscriber_count': int(channel_info.get('statistics', {}).get('subscriberCount', 0))
                }
            
            channel_data[channel_id]['outliers'].append(result)
        
        # Calculate metrics for each channel
        for channel_id, data_item in channel_data.items():
            outliers = data_item['outliers']
            outlier_scores = [r.get('outlierScore', 0) for r in outliers]
            brand_fits = [r.get('brandFit', 0) for r in outliers]
            
            data_item.update({
                'outlier_count': len(outliers),
                'avg_outlier_score': sum(outlier_scores) / len(outlier_scores) if outlier_scores else 0,
                'avg_brand_fit': sum(brand_fits) / len(brand_fits) if brand_fits else 0,
                'total_views': sum(int(r.get('statistics', {}).get('viewCount', 0)) for r in outliers),
                'success_rate': len([s for s in outlier_scores if s > 20]) / len(outlier_scores) * 100 if outlier_scores else 0,
                'consistency_score': self._calculate_consistency(outlier_scores),
                'growth_potential': self._assess_growth_potential(data_item),
                'recommendation': self._generate_channel_recommendation(data_item)
            })
        
        return channel_data
    
    def _calculate_consistency(self, scores):
        if len(scores) < 2:
            return 'N/A'
        
        import statistics
        std_dev = statistics.stdev(scores)
        mean_score = statistics.mean(scores)
        
        coefficient_of_variation = (std_dev / mean_score) if mean_score > 0 else 1
        
        if coefficient_of_variation < 0.2:
            return 'Very Consistent'
        elif coefficient_of_variation < 0.4:
            return 'Consistent'
        elif coefficient_of_variation < 0.6:
            return 'Moderate'
        else:
            return 'Inconsistent'
    
    def _assess_growth_potential(self, channel_data):
        avg_score = channel_data['avg_outlier_score']
        outlier_count = channel_data['outlier_count']
        subscriber_count = channel_data['subscriber_count']
        
        if avg_score > 25 and outlier_count > 3:
            return 'High'
        elif avg_score > 20 and outlier_count > 2:
            return 'Medium'
        else:
            return 'Low'
    
    def _generate_channel_recommendation(self, channel_data):
        growth_potential = channel_data.get('growth_potential', 'Low')
        avg_score = channel_data['avg_outlier_score']
        
        if growth_potential == 'High':
            return 'Priority collaboration target'
        elif growth_potential == 'Medium':
            return 'Monitor for future opportunities'
        else:
            return 'Consider alternative channels'
    
    def _generate_recommendations(self, data):
        recommendations = {
            'Content Strategy': [
                'Focus on gaming content, which shows highest outlier potential',
                'Create content similar to top-performing videos identified',
                'Target 10-15 minute video length for optimal engagement'
            ],
            'Channel Partnerships': [
                'Collaborate with channels showing consistent outlier performance',
                'Consider guest appearances on channels with 10K-100K subscribers',
                'Explore cross-promotion opportunities with complementary content creators'
            ],
            'Content Optimization': [
                'Use trending tags and keywords from successful outliers',
                'Optimize upload timing based on successful video patterns',
                'Implement thumbnail and title strategies from top performers'
            ],
            'Growth Opportunities': [
                'Analyze content gaps in high-performing niches',
                'Consider expanding into underserved content categories',
                'Monitor emerging trends for early adoption opportunities'
            ]
        }
        
        return recommendations
    
    def _calculate_statistics(self, data):
        results = data['results']
        
        if not results:
            return {}
        
        outlier_scores = [r.get('outlierScore', 0) for r in results]
        brand_fits = [r.get('brandFit', 0) for r in results]
        views = [int(r.get('statistics', {}).get('viewCount', 0)) for r in results]
        
        import statistics
        
        return {
            'Total Videos Analyzed': len(results),
            'Mean Outlier Score': round(statistics.mean(outlier_scores), 2),
            'Median Outlier Score': round(statistics.median(outlier_scores), 2),
            'Outlier Score Std Dev': round(statistics.stdev(outlier_scores) if len(outlier_scores) > 1 else 0, 2),
            'Mean Brand Fit': round(statistics.mean(brand_fits), 2),
            'Total Views (All Videos)': sum(views),
            'Average Views per Video': round(sum(views) / len(views), 0),
            'High Performance Videos (>25 score)': len([s for s in outlier_scores if s > 25]),
            'Brand Compatible Videos (>7 fit)': len([b for b in brand_fits if b > 7])
        }

# Global instance
enhanced_export_service = EnhancedExportService()
