import os
import io
import json
import uuid
import csv
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from urllib.parse import quote

# Excel and data processing
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    import pandas as pd
except ImportError:
    openpyxl = None
    pd = None

# PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.barcharts import VerticalBarChart
except ImportError:
    pass

# Charting
try:
    import matplotlib.pyplot as plt
    import seaborn as sns
    import matplotlib
    matplotlib.use('Agg')  # Use non-interactive backend
except ImportError:
    plt = None
    sns = None

from utils.logger import logger
from db import redis_client

class ExportService:
    """Service for exporting analysis results in various formats"""
    
    def __init__(self):
        self.temp_dir = os.path.join(os.getcwd(), 'temp')
        os.makedirs(self.temp_dir, exist_ok=True)
        
        # Configure matplotlib and seaborn for better charts
        if plt and sns:
            plt.style.use('seaborn-v0_8')
            sns.set_palette("husl")
    
    def get_analysis_data(self, analysis_id: str) -> Optional[Dict[str, Any]]:
        """Retrieve analysis data from Redis"""
        try:
            # Get analysis metadata
            analysis_data = redis_client.get(f'analysis:{analysis_id}')
            if not analysis_data:
                return None
            
            analysis = json.loads(analysis_data)
            
            # Get results
            results_data = redis_client.get(f'analysis_results:{analysis_id}')
            results = json.loads(results_data) if results_data else []
            
            return {
                'metadata': analysis,
                'results': results,
                'summary': analysis.get('summary', {})
            }
        except Exception as e:
            logger.error(f"Error retrieving analysis data for {analysis_id}: {e}")
            return None
    
    def export_to_excel(self, analysis_id: str, user_id: str = None) -> io.BytesIO:
        """Export analysis results to Excel format with multiple sheets"""
        if not openpyxl or not pd:
            raise ValueError("Excel export requires openpyxl and pandas")
        
        data = self.get_analysis_data(analysis_id)
        if not data:
            raise ValueError(f"Analysis {analysis_id} not found")
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default worksheet
        wb.remove(wb.active)
        
        # Create Summary sheet
        self._create_summary_sheet(wb, data)
        
        # Create Outlier Results sheet
        self._create_outliers_sheet(wb, data)
        
        # Create Channel Analysis sheet
        self._create_channels_sheet(wb, data)
        
        # Create Performance Charts sheet
        self._create_charts_sheet(wb, data)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Excel export completed for analysis {analysis_id}")
        return output
    
    def _create_summary_sheet(self, wb, data: Dict[str, Any]):
        """Create summary sheet with key metrics"""
        ws = wb.create_sheet("Summary")
        
        # Header styling
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        
        # Title
        ws['A1'] = 'YouTube Outlier Discovery Analysis'
        ws['A1'].font = Font(name='Arial', size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Analysis info
        metadata = data['metadata']
        summary = data['summary']
        
        row = 3
        info_data = [
            ('Analysis ID', metadata.get('id', 'N/A')),
            ('Created At', metadata.get('created_at', 'N/A')),
            ('Status', metadata.get('status', 'N/A')),
            ('Total Outliers Found', summary.get('totalOutliers', 0)),
            ('Channels Analyzed', summary.get('channelsAnalyzed', 0)),
            ('Exclusion Games', summary.get('exclusionGames', 0)),
            ('', ''),  # Spacer
            ('Configuration', ''),
            ('Min Subscribers', metadata.get('config', {}).get('minSubs', 'N/A')),
            ('Max Subscribers', metadata.get('config', {}).get('maxSubs', 'N/A')),
            ('Time Window (days)', metadata.get('config', {}).get('timeWindow', 'N/A')),
            ('Outlier Threshold', metadata.get('config', {}).get('outlierThreshold', 'N/A')),
            ('Brand Fit Threshold', metadata.get('config', {}).get('brandFitThreshold', 'N/A'))
        ]
        
        for label, value in info_data:
            if label:  # Skip empty spacer rows
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                
                if label == 'Configuration':  # Section header
                    ws[f'A{row}'].font = Font(bold=True)
                    ws[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            row += 1
        
        # Top outliers table
        ws[f'A{row + 1}'] = 'Top 10 Outliers'
        ws[f'A{row + 1}'].font = Font(bold=True, size=12)
        row += 3
        
        # Headers for top outliers
        headers = ['Video Title', 'Channel', 'Outlier Score', 'Brand Fit', 'Views', 'Subscribers']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Top outliers data
        results = data['results'][:10]  # Top 10
        for i, result in enumerate(results, 1):
            ws.cell(row=row + i, column=1, value=result.get('snippet', {}).get('title', 'N/A'))
            ws.cell(row=row + i, column=2, value=result.get('channelInfo', {}).get('snippet', {}).get('title', 'N/A'))
            ws.cell(row=row + i, column=3, value=round(result.get('outlierScore', 0), 2))
            ws.cell(row=row + i, column=4, value=round(result.get('brandFit', 0), 2))
            ws.cell(row=row + i, column=5, value=int(result.get('statistics', {}).get('viewCount', 0)))
            ws.cell(row=row + i, column=6, value=int(result.get('channelInfo', {}).get('statistics', {}).get('subscriberCount', 0)))
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_outliers_sheet(self, wb, data: Dict[str, Any]):
        """Create detailed outliers sheet"""
        ws = wb.create_sheet("Outlier Results")
        
        # Headers
        headers = [
            'Video Title', 'Channel Name', 'Video URL', 'Channel URL',
            'Outlier Score', 'Brand Fit Score', 'Views', 'Likes', 'Comments',
            'Subscriber Count', 'Video Count', 'Published Date', 'Duration',
            'Tags', 'Description Preview'
        ]
        
        # Style headers
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        results = data['results']
        for row, result in enumerate(results, 2):
            snippet = result.get('snippet', {})
            stats = result.get('statistics', {})
            channel_info = result.get('channelInfo', {})
            channel_snippet = channel_info.get('snippet', {})
            channel_stats = channel_info.get('statistics', {})
            
            # Video URL
            video_id = result.get('id', {}).get('videoId', result.get('id', ''))
            video_url = f"https://www.youtube.com/watch?v={video_id}" if video_id else 'N/A'
            
            # Channel URL
            channel_id = channel_info.get('id', '')
            channel_url = f"https://www.youtube.com/channel/{channel_id}" if channel_id else 'N/A'
            
            # Description preview (first 100 chars)
            description = snippet.get('description', '')
            description_preview = description[:100] + '...' if len(description) > 100 else description
            
            # Tags
            tags = ', '.join(snippet.get('tags', [])[:5])  # First 5 tags
            
            row_data = [
                snippet.get('title', 'N/A'),
                channel_snippet.get('title', 'N/A'),
                video_url,
                channel_url,
                round(result.get('outlierScore', 0), 2),
                round(result.get('brandFit', 0), 2),
                int(stats.get('viewCount', 0)),
                int(stats.get('likeCount', 0)),
                int(stats.get('commentCount', 0)),
                int(channel_stats.get('subscriberCount', 0)),
                int(channel_stats.get('videoCount', 0)),
                snippet.get('publishedAt', 'N/A'),
                snippet.get('duration', 'N/A'),
                tags,
                description_preview
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                
                # Add hyperlinks for URLs
                if col in [3, 4] and value.startswith('http'):
                    cell.hyperlink = value
                    cell.font = Font(color='0000FF', underline='single')
        
        # Conditional formatting for outlier scores
        from openpyxl.formatting.rule import ColorScaleRule
        outlier_score_col = 'E2:E' + str(len(results) + 1)
        ws.conditional_formatting.add(outlier_score_col,
                                    ColorScaleRule(start_type='min', start_color='FFFF00',
                                                 end_type='max', end_color='FF0000'))
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_channels_sheet(self, wb, data: Dict[str, Any]):
        """Create channel analysis sheet"""
        ws = wb.create_sheet("Channel Analysis")
        
        # Group results by channel
        channel_data = {}
        for result in data['results']:
            channel_info = result.get('channelInfo', {})
            channel_id = channel_info.get('id', 'unknown')
            
            if channel_id not in channel_data:
                channel_data[channel_id] = {
                    'info': channel_info,
                    'outliers': [],
                    'total_outlier_score': 0,
                    'avg_brand_fit': 0
                }
            
            channel_data[channel_id]['outliers'].append(result)
            channel_data[channel_id]['total_outlier_score'] += result.get('outlierScore', 0)
        
        # Calculate averages
        for channel_id, data_item in channel_data.items():
            outlier_count = len(data_item['outliers'])
            if outlier_count > 0:
                data_item['avg_outlier_score'] = data_item['total_outlier_score'] / outlier_count
                data_item['avg_brand_fit'] = sum(r.get('brandFit', 0) for r in data_item['outliers']) / outlier_count
        
        # Headers
        headers = [
            'Channel Name', 'Channel URL', 'Subscribers', 'Total Videos',
            'Outliers Found', 'Avg Outlier Score', 'Avg Brand Fit',
            'Total Views (Outliers)', 'Channel Description'
        ]
        
        # Style headers
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        sorted_channels = sorted(channel_data.items(), 
                               key=lambda x: x[1]['avg_outlier_score'], 
                               reverse=True)
        
        for row, (channel_id, ch_data) in enumerate(sorted_channels, 2):
            info = ch_data['info']
            snippet = info.get('snippet', {})
            stats = info.get('statistics', {})
            
            channel_url = f"https://www.youtube.com/channel/{channel_id}"
            total_outlier_views = sum(int(r.get('statistics', {}).get('viewCount', 0)) 
                                    for r in ch_data['outliers'])
            
            description = snippet.get('description', '')
            description_preview = description[:200] + '...' if len(description) > 200 else description
            
            row_data = [
                snippet.get('title', 'N/A'),
                channel_url,
                int(stats.get('subscriberCount', 0)),
                int(stats.get('videoCount', 0)),
                len(ch_data['outliers']),
                round(ch_data['avg_outlier_score'], 2),
                round(ch_data['avg_brand_fit'], 2),
                total_outlier_views,
                description_preview
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                
                # Add hyperlink for channel URL
                if col == 2 and value.startswith('http'):
                    cell.hyperlink = value
                    cell.font = Font(color='0000FF', underline='single')
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_charts_sheet(self, wb, data: Dict[str, Any]):
        """Create charts sheet with visualizations"""
        ws = wb.create_sheet("Performance Charts")
        
        # Prepare data for charts
        results = data['results'][:20]  # Top 20 for charts
        
        # Chart 1: Outlier Score Distribution
        ws['A1'] = 'Top 20 Videos by Outlier Score'
        ws['A1'].font = Font(bold=True, size=14)
        
        # Data for outlier score chart
        ws['A3'] = 'Video Title'
        ws['B3'] = 'Outlier Score'
        
        for i, result in enumerate(results, 4):
            title = result.get('snippet', {}).get('title', 'N/A')
            # Truncate long titles
            if len(title) > 30:
                title = title[:30] + '...'
            ws[f'A{i}'] = title
            ws[f'B{i}'] = result.get('outlierScore', 0)
        
        # Create bar chart for outlier scores
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Top Videos by Outlier Score"
        chart1.y_axis.title = 'Outlier Score'
        chart1.x_axis.title = 'Videos'
        
        data_ref = Reference(ws, min_col=2, min_row=3, max_row=len(results) + 3)
        cats_ref = Reference(ws, min_col=1, min_row=4, max_row=len(results) + 3)
        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(cats_ref)
        
        ws.add_chart(chart1, "D3")
        
        # Chart 2: Brand Fit vs Outlier Score Scatter
        start_row = len(results) + 8
        ws[f'A{start_row}'] = 'Brand Fit vs Outlier Score Analysis'
        ws[f'A{start_row}'].font = Font(bold=True, size=14)
        
        ws[f'A{start_row + 2}'] = 'Outlier Score'
        ws[f'B{start_row + 2}'] = 'Brand Fit Score'
        
        for i, result in enumerate(results, start_row + 3):
            ws[f'A{i}'] = result.get('outlierScore', 0)
            ws[f'B{i}'] = result.get('brandFit', 0)
        
        # Create scatter chart
        from openpyxl.chart import ScatterChart
        chart2 = ScatterChart()
        chart2.title = "Brand Fit vs Outlier Score"
        chart2.style = 13
        chart2.x_axis.title = 'Outlier Score'
        chart2.y_axis.title = 'Brand Fit Score'
        
        xvalues = Reference(ws, min_col=1, min_row=start_row + 3, max_row=start_row + len(results) + 2)
        yvalues = Reference(ws, min_col=2, min_row=start_row + 3, max_row=start_row + len(results) + 2)
        series = openpyxl.chart.Series(yvalues, xvalues, title="Videos")
        chart2.series.append(series)
        
        ws.add_chart(chart2, f"D{start_row + 2}")
    
    def export_to_csv(self, analysis_id: str, user_id: str = None) -> io.StringIO:
        """Export analysis results to CSV format"""
        data = self.get_analysis_data(analysis_id)
        if not data:
            raise ValueError(f"Analysis {analysis_id} not found")
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            'Video Title', 'Channel Name', 'Video URL', 'Channel URL',
            'Outlier Score', 'Brand Fit Score', 'Views', 'Likes', 'Comments',
            'Subscriber Count', 'Video Count', 'Published Date', 'Duration'
        ]
        writer.writerow(headers)
        
        # Data rows
        for result in data['results']:
            snippet = result.get('snippet', {})
            stats = result.get('statistics', {})
            channel_info = result.get('channelInfo', {})
            channel_snippet = channel_info.get('snippet', {})
            channel_stats = channel_info.get('statistics', {})
            
            video_id = result.get('id', {}).get('videoId', result.get('id', ''))
            video_url = f"https://www.youtube.com/watch?v={video_id}" if video_id else 'N/A'
            
            channel_id = channel_info.get('id', '')
            channel_url = f"https://www.youtube.com/channel/{channel_id}" if channel_id else 'N/A'
            
            row = [
                snippet.get('title', 'N/A'),
                channel_snippet.get('title', 'N/A'),
                video_url,
                channel_url,
                round(result.get('outlierScore', 0), 2),
                round(result.get('brandFit', 0), 2),
                int(stats.get('viewCount', 0)),
                int(stats.get('likeCount', 0)),
                int(stats.get('commentCount', 0)),
                int(channel_stats.get('subscriberCount', 0)),
                int(channel_stats.get('videoCount', 0)),
                snippet.get('publishedAt', 'N/A'),
                snippet.get('duration', 'N/A')
            ]
            writer.writerow(row)
        
        output.seek(0)
        logger.info(f"CSV export completed for analysis {analysis_id}")
        return output
    
    def export_to_json(self, analysis_id: str, user_id: str = None) -> str:
        """Export analysis results to JSON format"""
        data = self.get_analysis_data(analysis_id)
        if not data:
            raise ValueError(f"Analysis {analysis_id} not found")
        
        # Clean up data for JSON export
        export_data = {
            'analysis_id': analysis_id,
            'exported_at': datetime.utcnow().isoformat(),
            'metadata': data['metadata'],
            'summary': data['summary'],
            'results': data['results']
        }
        
        logger.info(f"JSON export completed for analysis {analysis_id}")
        return json.dumps(export_data, indent=2, default=str)
    
    def export_to_pdf(self, analysis_id: str, user_id: str = None) -> io.BytesIO:
        """Export analysis results to PDF format"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
        except ImportError:
            raise ValueError("PDF export requires reportlab")
        
        data = self.get_analysis_data(analysis_id)
        if not data:
            raise ValueError(f"Analysis {analysis_id} not found")
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        heading_style = styles['Heading1']
        normal_style = styles['Normal']
        
        # Content
        story = []
        
        # Title
        story.append(Paragraph("YouTube Outlier Discovery Analysis Report", title_style))
        story.append(Spacer(1, 12))
        
        # Summary
        metadata = data['metadata']
        summary = data['summary']
        
        story.append(Paragraph("Analysis Summary", heading_style))
        summary_data = [
            ['Analysis ID', metadata.get('id', 'N/A')],
            ['Created At', metadata.get('created_at', 'N/A')],
            ['Status', metadata.get('status', 'N/A')],
            ['Total Outliers Found', str(summary.get('totalOutliers', 0))],
            ['Channels Analyzed', str(summary.get('channelsAnalyzed', 0))],
            ['Exclusion Games', str(summary.get('exclusionGames', 0))]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 12))
        
        # Top 20 Outliers
        story.append(Paragraph("Top 20 Outlier Videos", heading_style))
        
        results_data = [['Video Title', 'Channel', 'Outlier Score', 'Brand Fit', 'Views']]
        for result in data['results'][:20]:
            title = result.get('snippet', {}).get('title', 'N/A')
            if len(title) > 50:
                title = title[:50] + '...'
            
            results_data.append([
                title,
                result.get('channelInfo', {}).get('snippet', {}).get('title', 'N/A')[:30],
                str(round(result.get('outlierScore', 0), 2)),
                str(round(result.get('brandFit', 0), 2)),
                str(int(result.get('statistics', {}).get('viewCount', 0)))
            ])
        
        results_table = Table(results_data)
        results_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8)
        ]))
        story.append(results_table)
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        logger.info(f"PDF export completed for analysis {analysis_id}")
        return output
    
    def export_to_html(self, analysis_id: str, user_id: str = None) -> str:
        """Export analysis results to HTML format"""
        try:
            from jinja2 import Template
        except ImportError:
            raise ValueError("HTML export requires jinja2")
        
        data = self.get_analysis_data(analysis_id)
        if not data:
            raise ValueError(f"Analysis {analysis_id} not found")
        
        html_template = Template("""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>YouTube Outlier Discovery Analysis Report</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        h1 {
            color: #2E75B6;
            text-align: center;
            margin-bottom: 30px;
        }
        h2 {
            color: #333;
            border-bottom: 2px solid #2E75B6;
            padding-bottom: 5px;
        }
        .summary {
            background-color: #f8f9fa;
            padding: 15px;
            border-radius: 5px;
            margin-bottom: 20px;
        }
        .summary-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 10px;
        }
        .summary-item {
            background-color: white;
            padding: 10px;
            border-radius: 5px;
            border-left: 4px solid #2E75B6;
        }
        .summary-label {
            font-weight: bold;
            color: #666;
            font-size: 0.9em;
        }
        .summary-value {
            font-size: 1.2em;
            color: #333;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
        }
        th, td {
            padding: 10px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }
        th {
            background-color: #2E75B6;
            color: white;
            font-weight: bold;
        }
        tr:nth-child(even) {
            background-color: #f8f9fa;
        }
        tr:hover {
            background-color: #e8f4fd;
        }
        .outlier-score {
            font-weight: bold;
            color: #d63384;
        }
        .brand-fit {
            font-weight: bold;
            color: #198754;
        }
        .video-link {
            color: #2E75B6;
            text-decoration: none;
        }
        .video-link:hover {
            text-decoration: underline;
        }
        .footer {
            text-align: center;
            margin-top: 30px;
            color: #666;
            font-size: 0.9em;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>YouTube Outlier Discovery Analysis Report</h1>
        
        <div class="summary">
            <h2>Analysis Summary</h2>
            <div class="summary-grid">
                <div class="summary-item">
                    <div class="summary-label">Analysis ID</div>
                    <div class="summary-value">{{ metadata.id }}</div>
                </div>
                <div class="summary-item">
                    <div class="summary-label">Created At</div>
                    <div class="summary-value">{{ metadata.created_at }}</div>
                </div>
                <div class="summary-item">
                    <div class="summary-label">Total Outliers</div>
                    <div class="summary-value">{{ summary.totalOutliers }}</div>
                </div>
                <div class="summary-item">
                    <div class="summary-label">Channels Analyzed</div>
                    <div class="summary-value">{{ summary.channelsAnalyzed }}</div>
                </div>
                <div class="summary-item">
                    <div class="summary-label">Outlier Threshold</div>
                    <div class="summary-value">{{ metadata.config.outlierThreshold }}</div>
                </div>
                <div class="summary-item">
                    <div class="summary-label">Brand Fit Threshold</div>
                    <div class="summary-value">{{ metadata.config.brandFitThreshold }}</div>
                </div>
            </div>
        </div>
        
        <h2>Outlier Videos (Top {{ results|length }})</h2>
        <table>
            <thead>
                <tr>
                    <th>Video Title</th>
                    <th>Channel</th>
                    <th>Outlier Score</th>
                    <th>Brand Fit</th>
                    <th>Views</th>
                    <th>Subscribers</th>
                    <th>Published</th>
                </tr>
            </thead>
            <tbody>
                {% for result in results %}
                <tr>
                    <td>
                        <a href="https://www.youtube.com/watch?v={{ result.id.videoId or result.id }}" 
                           class="video-link" target="_blank">
                            {{ result.snippet.title }}
                        </a>
                    </td>
                    <td>
                        <a href="https://www.youtube.com/channel/{{ result.channelInfo.id }}" 
                           class="video-link" target="_blank">
                            {{ result.channelInfo.snippet.title }}
                        </a>
                    </td>
                    <td class="outlier-score">{{ "%.2f"|format(result.outlierScore) }}</td>
                    <td class="brand-fit">{{ "%.2f"|format(result.brandFit) }}</td>
                    <td>{{ "{:,}".format(result.statistics.viewCount|int) }}</td>
                    <td>{{ "{:,}".format(result.channelInfo.statistics.subscriberCount|int) }}</td>
                    <td>{{ result.snippet.publishedAt[:10] }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
        
        <div class="footer">
            <p>Generated by YouTube Outlier Discovery Tool on {{ exported_at }}</p>
        </div>
    </div>
</body>
</html>
        """)
        
        html_content = html_template.render(
            metadata=data['metadata'],
            summary=data['summary'],
            results=data['results'],
            exported_at=datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
        )
        
        logger.info(f"HTML export completed for analysis {analysis_id}")
        return html_content
    
    def create_export_job(self, analysis_id: str, format_type: str, user_id: str = None) -> str:
        """Create a background export job for large files"""
        job_id = str(uuid.uuid4())
        
        job_data = {
            'id': job_id,
            'analysis_id': analysis_id,
            'format': format_type,
            'user_id': user_id,
            'status': 'pending',
            'created_at': datetime.utcnow().isoformat(),
            'progress': 0
        }
        
        # Store job in Redis
        redis_client.setex(f'export_job:{job_id}', 3600, json.dumps(job_data))
        
        # In a real implementation, this would be queued with Celery or similar
        # For now, we'll process it immediately
        import threading
        
        def process_export():
            try:
                job_data['status'] = 'processing'
                job_data['progress'] = 50
                redis_client.setex(f'export_job:{job_id}', 3600, json.dumps(job_data))
                
                # Generate export
                if format_type == 'excel':
                    result = self.export_to_excel(analysis_id, user_id)
                elif format_type == 'csv':
                    result = self.export_to_csv(analysis_id, user_id)
                elif format_type == 'pdf':
                    result = self.export_to_pdf(analysis_id, user_id)
                elif format_type == 'json':
                    result = self.export_to_json(analysis_id, user_id)
                elif format_type == 'html':
                    result = self.export_to_html(analysis_id, user_id)
                else:
                    raise ValueError(f"Unsupported format: {format_type}")
                
                # Save to temporary file
                filename = f"analysis_{analysis_id}_{job_id}.{format_type}"
                filepath = os.path.join(self.temp_dir, filename)
                
                if isinstance(result, (io.BytesIO, io.StringIO)):
                    mode = 'wb' if isinstance(result, io.BytesIO) else 'w'
                    with open(filepath, mode) as f:
                        f.write(result.getvalue())
                else:
                    with open(filepath, 'w', encoding='utf-8') as f:
                        f.write(result)
                
                job_data['status'] = 'completed'
                job_data['progress'] = 100
                job_data['file_path'] = filepath
                job_data['filename'] = filename
                redis_client.setex(f'export_job:{job_id}', 3600, json.dumps(job_data))
                
            except Exception as e:
                logger.error(f"Export job {job_id} failed: {e}")
                job_data['status'] = 'failed'
                job_data['error'] = str(e)
                redis_client.setex(f'export_job:{job_id}', 3600, json.dumps(job_data))
        
        thread = threading.Thread(target=process_export)
        thread.start()
        
        logger.info(f"Export job {job_id} created for analysis {analysis_id}, format {format_type}")
        return job_id
    
    def get_export_job_status(self, job_id: str) -> Optional[Dict[str, Any]]:
        """Get export job status"""
        job_data = redis_client.get(f'export_job:{job_id}')
        if job_data:
            return json.loads(job_data)
        return None
    
    def cleanup_temp_files(self, max_age_hours: int = 24):
        """Clean up old temporary export files"""
        cutoff_time = datetime.utcnow() - timedelta(hours=max_age_hours)
        cleaned_count = 0
        
        try:
            for filename in os.listdir(self.temp_dir):
                filepath = os.path.join(self.temp_dir, filename)
                if os.path.isfile(filepath):
                    file_time = datetime.fromtimestamp(os.path.getctime(filepath))
                    if file_time < cutoff_time:
                        try:
                            os.remove(filepath)
                            cleaned_count += 1
                            logger.info(f"Cleaned up old export file: {filename}")
                        except Exception as e:
                            logger.error(f"Error cleaning up file {filename}: {e}")
            
            # Also clean up export jobs from Redis
            try:
                from .export_queue import get_export_queue_manager
                queue_manager = get_export_queue_manager()
                cleaned_count += queue_manager.cleanup_old_jobs(max_age_hours)
            except Exception as e:
                logger.warning(f"Could not clean up export jobs: {e}")
                
        except Exception as e:
            logger.error(f"Error during cleanup: {e}")
        
        return cleaned_count
